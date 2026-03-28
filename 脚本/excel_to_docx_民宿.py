# -*- coding: utf-8 -*-
"""
将“美食.xlsx”中的每一行数据生成为一个 .docx 文档。

适用场景：
1. Excel 中每一行代表一家饭店；
2. 需要把每一行整理成固定模板样式的 Word 文档；
3. 文档内容结构参考“美食模板.docx”的版式。

使用方法：
    pip install openpyxl python-docx
    python excel_to_docx_美食.py

说明：
- 默认读取当前目录下的“美食.xlsx”；
- 默认在当前目录下生成“生成结果”文件夹；
- 每一行生成一个 docx 文件；
- 文件名默认使用：序号_饭店名称.docx
"""

import os
import re
import copy
from typing import Any

from openpyxl import load_workbook
from docx import Document


# =========================
# 这里是可按需修改的配置
# =========================
EXCEL_PATH = "test.xlsx"          # Excel 文件路径
TEMPLATE_PATH = r"模板\民宿模板.docx"   # Word 模板文件路径
OUTPUT_DIR = "生成结果/民宿"           # 输出文件夹
SHEET_NAME = "民宿"                 # 读取的工作表名；如果为 None，则默认读取第一个表


def safe_str(value: Any) -> str:
    """把单元格值安全地转成字符串。空值转为'无'。"""
    if value is None:
        return "无"
    val_str = str(value).strip()
    return val_str if val_str else "无"


def sanitize_filename(name: str) -> str:
    """清理文件名中的非法字符，避免 Windows 下保存失败。"""
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    return name.strip() or "未命名"


def replace_placeholders(doc: Document, row_data: dict) -> None:
    """
    遍历文档中的段落，将 {列名} 替换为 row_data 中对应的数据。
    原汁原味地保留模板格式（基于段落的首个 Run 样式）。
    """
    # 构建替换映射，Excel 里的列名对应占位符格式，比如 "饭店名称" 对应 "{饭店名称}"
    mapping = {}
    for key, value in row_data.items():
        mapping[f"{{{key}}}"] = safe_str(value)

    # --- 增加字段别名兼容处理 ---
    # 解决 Excel 表头和 Word 模板占位符名字对不上的问题
    mapping["{民宿名称}"] = mapping.get("{酒店名称}", "无")
    mapping["{电话号码}"] = mapping.get("{酒店电话}", "无")
    mapping["{联系人}"] = mapping.get("{联系人身份及称谓}", "无")
    mapping["{各个房间价格及其数目}"] = mapping.get("{各个房间类型的价格及其数量}", "无")
    mapping["{前台服务}"] = mapping.get("{前台服务有哪些}", "无")
    mapping["{游客评价}"] = mapping.get("{顾客评价}", "无")
    mapping["{注意事项}"] = mapping.get("{是否需要提前预约（若需要，需要提前几天）能否提前退订}", "无")

    # 遍历并替换所有段落文本
    for p in doc.paragraphs:
        if not p.text.strip():
            continue
            
        original_text = p.text
        new_text = original_text
        
        # 查找是否存在需要替换的占位符
        for placeholder, actual_val in mapping.items():
            if placeholder in new_text:
                new_text = new_text.replace(placeholder, actual_val)
                
        # 如果文本发生了实际替换变化
        if new_text != original_text and p.runs:
            # 针对包含多行的文本（如特色菜、简介），去除首尾空白并按换行符拆分，过滤掉空行
            lines = [line.strip() for line in new_text.split('\n') if line.strip()]
            
            # 如果内容完全为空了（比如该饭店没有特色菜），删除这个段落，不留空行
            if not lines:
                p._element.getparent().remove(p._element)
                continue
                
            # 将第一行写入当前段落的首个 run，清空其他 run，保障不会破坏段落原本的格式
            for i, run in enumerate(p.runs):
                if i == 0:
                    run.text = lines[0]
                else:
                    run.text = ""

            # 如果内容被拆分出了多行，为它们逐一创建复制的新段落，以确保100%保留原有的格式（含缩进、字体、颜色等）
            curr_element = p._element
            from docx.text.paragraph import Paragraph
            for line in lines[1:]:
                # 深度拷贝原段落的底层 XML 元素，完美保留所有一切样式和格式
                new_p_elem = copy.deepcopy(p._element)
                curr_element.addnext(new_p_elem)
                
                # 包装为 Paragraph 对象方便操作文本
                new_p = Paragraph(new_p_elem, p._parent)
                for i, run in enumerate(new_p.runs):
                    if i == 0:
                        run.text = line
                    else:
                        run.text = ""
                        
                curr_element = new_p_elem


def build_doc_from_row(row_data: dict, output_path: str) -> None:
    """根据一行 Excel 数据，使用模板生成一个 Word 文档。"""
    # 加载带有格式的模板文档
    doc = Document(TEMPLATE_PATH)
    
    # 执行内容替换
    replace_placeholders(doc, row_data)
    
    # 保存替换好内容的新文档
    doc.save(output_path)


def main() -> None:
    """主函数：读取 Excel，逐行生成 docx。"""
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"未找到 Excel 文件：{EXCEL_PATH}")
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"未找到 Word 模板文件：{TEMPLATE_PATH}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]

    # 读取表头（第一行）
    headers = [safe_str(cell.value) for cell in ws[1]]

    # 遍历数据行（从第二行开始）
    generated_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过整行为空的记录
        if all(cell is None or safe_str(cell) == '' for cell in row):
            continue

        # 把本行数据组装成字典：{列名: 单元格值}
        row_data = {headers[i]: row[i] for i in range(len(headers))}

        # 读取序号和饭店名称，用于生成文件名
        seq = safe_str(row_data.get('序号', ''))
        shop_name = safe_str(row_data.get('酒店名称', ''))

        filename_parts = []
        if seq:
            filename_parts.append(seq)
        if shop_name:
            filename_parts.append(shop_name)

        if not filename_parts:
            filename_parts.append(f"第{row_idx}行")

        filename = sanitize_filename('_'.join(filename_parts) + '.docx')
        output_path = os.path.join(OUTPUT_DIR, filename)

        try:
            build_doc_from_row(row_data, output_path)
            generated_count += 1
            print(f"已生成：{output_path}")
        except PermissionError:
            print(f"【报错】无法保存 {output_path}，文件可能正在被其他程序（如 Word）打开，请关闭后重试！")
        except Exception as e:
            print(f"【报错】生成 {output_path} 失败：{e}")

    print(f"\n全部完成，共生成 {generated_count} 个 Word 文档。")
    print(f"输出目录：{os.path.abspath(OUTPUT_DIR)}")


if __name__ == '__main__':
    main()
